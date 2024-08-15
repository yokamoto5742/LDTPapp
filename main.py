import os
import re

from flet_core import dropdown, Dropdown
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from datetime import datetime
import csv

import flet as ft
import pandas as pd
from flet import View
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from sqlalchemy import create_engine, Column, Integer, String, Float, Date, Boolean
from sqlalchemy.orm import declarative_base
from sqlalchemy.orm import sessionmaker
import configparser
from contextlib import contextmanager
import threading

from barcode.codex import Code128
from barcode.writer import ImageWriter
from io import BytesIO

VERSION = "1.0.0"
LAST_UPDATED = "2024/08/14"

# config.iniファイルの読み込み
config = configparser.ConfigParser()
config.read('config.ini', encoding='utf-8')
db_url = config.get('Database', 'db_url')
barcode_config = config['Barcode']
table_width = config.getint('DataTable', 'width')
document_number = config.get('Document', 'document_number', fallback='39221')
csv_file_path = config.get('FilePaths', 'patient_data')
export_folder = config.get('FilePaths', 'export_folder')

# SQLAlchemyの設定
engine = create_engine(db_url, pool_pre_ping=True, pool_size=10)
Session = sessionmaker(bind=engine)
Base = declarative_base()

# 治療計画書の履歴の選択を空欄にする(初期値)
selected_row = None


# PatientInfoモデルの定義
class PatientInfo(Base):
    __tablename__ = 'patient_info'
    id = Column(Integer, primary_key=True)
    patient_id = Column(Integer)
    patient_name = Column(String)
    kana = Column(String)
    gender = Column(String)
    birthdate = Column(Date)
    issue_date = Column(Date)
    doctor_id = Column(Integer)
    doctor_name = Column(String)
    department = Column(String)
    department_id = Column(Integer)
    main_diagnosis = Column(String)
    sheet_name = Column(String)
    creation_count = Column(Integer)
    goal1 = Column(String)  # ①達成目標
    goal2 = Column(String)  # ②行動目標
    target_weight = Column(Float)
    diet = Column(String)
    exercise_prescription = Column(String)
    exercise_time = Column(String)
    exercise_frequency = Column(String)
    exercise_intensity = Column(String)
    daily_activity = Column(String)
    nonsmoker = Column(Boolean)
    smoking_cessation = Column(Boolean)
    other1 = Column(String)
    other2 = Column(String)


class MainDisease(Base):
    __tablename__ = "main_diseases"
    id = Column(Integer, primary_key=True)
    name = Column(String)  # 主病名


class SheetName(Base):
    __tablename__ = "sheet_names"
    id = Column(Integer, primary_key=True)
    main_disease_id = Column(Integer)
    name = Column(String)  # シート名


class Template(Base):
    __tablename__ = 'templates'
    id = Column(Integer, primary_key=True)
    main_disease = Column(String)
    sheet_name = Column(String)
    goal1 = Column(String)
    goal2 = Column(String)
    diet = Column(String)
    exercise_prescription = Column(String)
    exercise_time = Column(String)
    exercise_frequency = Column(String)
    exercise_intensity = Column(String)
    daily_activity = Column(String)
    other1 = Column(String)
    other2 = Column(String)


# テーブルの作成
Base.metadata.create_all(engine)


class TreatmentPlanGenerator:
    @staticmethod
    def generate_plan(patient_info, file_name):
        template_path = config.get("Paths", "template_path")
        output_path = config.get("Paths", "output_path")

        current_datetime = datetime.now().strftime("%Y%m%d%H%M%S")
        patient_id = str(patient_info.patient_id).zfill(9)
        document_number = "39221"
        department_id = str(patient_info.department_id).zfill(3)
        doctor_id = str(patient_info.doctor_id).zfill(5)

        new_file_name = f"{patient_id}{document_number}{department_id}{doctor_id}{current_datetime}.xlsm"

        file_path = os.path.join(output_path, new_file_name)
        workbook = load_workbook(template_path, keep_vba=True)
        common_sheet = workbook["共通情報"]
        plan_sheet = workbook["計画書"]
        TreatmentPlanGenerator.populate_common_sheet(common_sheet, patient_info)

        options = {
            'write_text': barcode_config.getboolean('write_text', False),
            'module_height': barcode_config.getfloat('module_height', 15),
            'module_width': barcode_config.getfloat('module_width', 0.25),
            'quiet_zone': barcode_config.getint('quiet_zone', 1),
        }

        # バーコードの生成
        barcode_data = f"{patient_id}{document_number}{department_id}{doctor_id}{current_datetime}"
        barcode = Code128(barcode_data, writer=ImageWriter())
        buffer = BytesIO()
        barcode.write(buffer, options=options)

        # バーコード画像の挿入
        img = Image(buffer)
        img.width = barcode_config.getint('image_width', 200)
        img.height = barcode_config.getint('image_height', 30)
        image_position = barcode_config.get('image_position', 'B2')
        plan_sheet.add_image(img, image_position)

        workbook.save(file_path)

        wb = load_workbook(file_path, read_only=False, keep_vba=True)
        ws_common = wb["共通情報"]
        ws_common.sheet_view.tabSelected = False
        ws_plan = wb["計画書"]
        ws_plan.sheet_view.tabSelected = True
        wb.active = ws_plan
        wb.save(file_path)
        os.startfile(file_path)

    @staticmethod
    def populate_common_sheet(common_sheet, patient_info):
        common_sheet["B2"] = patient_info.patient_id
        common_sheet["B3"] = patient_info.patient_name
        common_sheet["B4"] = patient_info.kana
        common_sheet["B5"] = patient_info.gender
        common_sheet["B6"] = patient_info.birthdate
        common_sheet["B7"] = patient_info.issue_date
        common_sheet["B8"] = patient_info.doctor_id
        common_sheet["B9"] = patient_info.doctor_name
        common_sheet["B10"] = patient_info.department_id  # 診療科IDを追加
        common_sheet["B11"] = patient_info.department
        common_sheet["B12"] = patient_info.main_diagnosis
        common_sheet["B13"] = patient_info.creation_count
        common_sheet["B14"] = patient_info.target_weight
        common_sheet["B15"] = patient_info.sheet_name
        common_sheet["B16"] = patient_info.goal1
        common_sheet["B17"] = patient_info.goal2
        common_sheet["B18"] = patient_info.diet
        common_sheet["B19"] = patient_info.exercise_prescription
        common_sheet["B20"] = patient_info.exercise_time
        common_sheet["B21"] = patient_info.exercise_frequency
        common_sheet["B22"] = patient_info.exercise_intensity
        common_sheet["B23"] = patient_info.daily_activity
        common_sheet["B24"] = patient_info.nonsmoker
        common_sheet["B25"] = patient_info.smoking_cessation
        common_sheet["B26"] = patient_info.other1
        common_sheet["B27"] = patient_info.other2


class TemplateManager:
    def __init__(self):
        self.templates = {}

    def get_template(self, main_disease, sheet_name):
        return self.templates.get((main_disease, sheet_name))


class MyHandler(FileSystemEventHandler):
    def __init__(self, page):
        self.page = page

    def on_deleted(self, event):
        if event.src_path == csv_file_path:
            self.page.window.close()


class DropdownItems:
    def __init__(self):
        self.items = {
            'exercise_prescription': ['ウォーキング', 'ストレッチ', '筋トレ'],
            'exercise_time': ['20分', '30分', '60分'],
            'exercise_frequency': ['毎日', '週に5日', '週に3日'],
            'exercise_intensity': ['息が弾む程度', 'ニコニコペース'],
            'daily_activity': ['5000歩', '8000歩', '10000歩']
        }

    def get_options(self, key):
        return [ft.dropdown.Option(option) for option in self.items.get(key, [])]

    def add_item(self, key, options):
        self.items[key] = options

    def create_dropdown(self, key, label, width, on_change=None):
        return ft.Dropdown(
            label=label,
            width=width,
            options=self.get_options(key),
            on_change=on_change
        )


def create_form_fields(dropdown_items):
    exercise_prescription = dropdown_items.create_dropdown('exercise_prescription', "運動の種類", 200)
    exercise_time = dropdown_items.create_dropdown('exercise_time', "時間", 200)
    exercise_frequency = dropdown_items.create_dropdown('exercise_frequency', "頻度", 300)
    exercise_intensity = dropdown_items.create_dropdown('exercise_intensity', "強度", 300)
    daily_activity = dropdown_items.create_dropdown('daily_activity', "日常生活の活動量増加", 400)

    def create_focus_handler(next_field):
        return lambda _: next_field.focus()

    exercise_prescription.on_change = create_focus_handler(exercise_time)
    exercise_time.on_change = create_focus_handler(exercise_frequency)
    exercise_frequency.on_change = create_focus_handler(exercise_intensity)
    exercise_intensity.on_change = create_focus_handler(daily_activity)
    # daily_activityのon_changeは設定しない

    return exercise_prescription, exercise_time, exercise_frequency, exercise_intensity, daily_activity


def start_file_monitoring(page):
    event_handler = MyHandler(page)
    observer = Observer()
    observer.schedule(event_handler, path=os.path.dirname(csv_file_path), recursive=False)
    observer.start()
    return observer


def check_file_exists(page):
    if not os.path.exists(csv_file_path):
        page.window.close()


def load_patient_data():
    global csv_file_path
    try:
        config_csv = configparser.ConfigParser()
        config_csv.read('config.ini')
        csv_file_path = config_csv.get('FilePaths', 'patient_data')

        date_columns = [0, 6]  # 0列目と6列目を日付として読み込む
        nrows = 3  # csvファイルで先頭3行のみ読み込む

        df = pd.read_csv(csv_file_path, encoding="shift_jis", header=None, parse_dates=date_columns, nrows=nrows)
        return "", df

    except (configparser.NoSectionError, configparser.NoOptionError):
        return "エラー: config.iniファイルに'FilePaths'セクションまたは'patient_data'キーが見つかりません。", None
    except Exception as e:
        return f"エラー: {str(e)}", None


@contextmanager
def get_session():
    session = Session()
    try:
        yield session
    finally:
        session.close()





def load_main_diseases():
    with get_session() as session:
        main_diseases = session.query(MainDisease).all()
        return [ft.dropdown.Option(str(disease.name)) for disease in main_diseases]


def load_sheet_names(main_disease=None):
    with get_session() as session:
        if main_disease:
            sheet_names = session.query(SheetName).filter(SheetName.main_disease_id == main_disease).all()
        else:
            sheet_names = session.query(SheetName).all()
        return [ft.dropdown.Option(str(sheet.name)) for sheet in sheet_names]


def format_date(date_str):
    if pd.isna(date_str):  # pd.isna()で欠損値かどうかを判定
        return ""
    return pd.to_datetime(date_str).strftime("%Y/%m/%d")


def initialize_database():
    Base.metadata.create_all(engine)


def create_ui(page):
    page.title = "生活習慣病療養計画書"
    page.window.width = config.getint('settings', 'window_width', fallback=1200)
    page.window.height = config.getint('settings', 'window_height', fallback=800)
    page.locale_configuration = ft.LocaleConfiguration(
        supported_locales=[
            ft.Locale("ja", "JP"),
            ft.Locale("en", "US")
        ],
        current_locale=ft.Locale("ja", "JP")
    )

    dropdown_items = DropdownItems()

    threading.Thread(target=initialize_database).start()

    # pat.csvの読み込み
    error_message, df_patients = load_patient_data()
    initial_patient_id = ""

    if error_message:
        print(error_message)
    else:
        if df_patients is not None and not df_patients.empty:
            initial_patient_id = int(df_patients.iloc[0, 2])

    # 初期データの挿入
    session = Session()
    if session.query(MainDisease).count() == 0:
        main_diseases = [
            MainDisease(id=1, name="高血圧症"),
            MainDisease(id=2, name="脂質異常症"),
            MainDisease(id=3, name="糖尿病")
        ]
        session.add_all(main_diseases)
        session.commit()

    if session.query(SheetName).count() == 0:
        sheet_names = [
            SheetName(main_disease_id=1, name="1_血圧130-80以下"),
            SheetName(main_disease_id=1, name="2_血圧140-90以下"),
            SheetName(main_disease_id=1, name="3_血圧140-90以下_歩行"),
            SheetName(main_disease_id=2, name="1_LDL120以下"),
            SheetName(main_disease_id=2, name="2_LDL100以下"),
            SheetName(main_disease_id=2, name="3_LDL70以下"),
            SheetName(main_disease_id=3, name="1_HbA1c７％"),
            SheetName(main_disease_id=3, name="2_HbA1c６％"),
            SheetName(main_disease_id=3, name="3_HbA1c８％"),
        ]
        session.add_all(sheet_names)
        session.commit()

    if session.query(Template).count() == 0:
        templates = [
            Template(main_disease="高血圧症", sheet_name="1_血圧130-80以下",
                     goal1="家庭血圧が測定でき、朝と就寝前のいずれかで130/80mmHg以下",
                     goal2="塩分を控えた食事と運動習慣を目標にする",
                     diet="塩分量を適正にする/食物繊維の摂取量を増やす/ゆっくり食べる/間食を減らす",
                     exercise_prescription="ウォーキング", exercise_time="30分以上",
                     exercise_frequency="１週間に２回以上",
                     exercise_intensity="少し汗をかく程度", daily_activity="1日5000歩以上",
                     other1="睡眠の確保１日７時間", other2="毎日の歩数の測定"),
            Template(main_disease="高血圧症", sheet_name="2_血圧140-90以下",
                     goal1="家庭血圧が測定でき、朝と就寝前のいずれかで140/90mmHg以下",
                     goal2="塩分を控えた食事と運動習慣を目標にする",
                     diet="塩分量を適正にする/食物繊維の摂取量を増やす/ゆっくり食べる/間食を減らす",
                     exercise_prescription="ストレッチ運動", exercise_time="30分以上",
                     exercise_frequency="１週間に２回以上",
                     exercise_intensity="少し汗をかく程度", daily_activity="ストレッチ運動を主に行う",
                     other1="睡眠の確保１日７時間", other2="毎日の歩数の測定"),
            Template(main_disease="高血圧症", sheet_name="3_血圧140-90以下_歩行",
                     goal1="家庭血圧が測定でき、朝と就寝前のいずれかで140/90mmHg以下",
                     goal2="塩分を控えた食事と運動習慣を目標にする",
                     diet="塩分量を適正にする/食物繊維の摂取量を増やす/ゆっくり食べる/間食を減らす",
                     exercise_prescription="ウォーキング", exercise_time="30分以上",
                     exercise_frequency="１週間に２回以上",
                     exercise_intensity="少し汗をかく程度", daily_activity="1日6000歩以上",
                     other1="睡眠の確保１日７時間", other2="毎日の歩数の測定"),
            Template(main_disease="脂質異常症", sheet_name="1_LDL120以下", goal1="LDLコレステロール＜120/TG＜150/HDL≧40",
                     goal2="毎日の有酸素運動と食習慣の改善",
                     diet="食事摂取量を適正にする/食物繊維の摂取量を増やす/ゆっくり食べる/間食を減らす",
                     exercise_prescription="ウォーキング", exercise_time="30分以上",
                     exercise_frequency="１週間に２回以上",
                     exercise_intensity="少し汗をかく程度", daily_activity="1日5000歩以上",
                     other1="飲酒の制限、肥満度の改善", other2="毎日の歩数の測定"),
            Template(main_disease="脂質異常症", sheet_name="2_LDL100以下", goal1="LDLコレステロール＜100/TG＜150/HDL≧40",
                     goal2="毎日の有酸素運動と食習慣の改善",
                     diet="食事摂取量を適正にする/食物繊維の摂取量を増やす/ゆっくり食べる/間食を減らす",
                     exercise_prescription="ウォーキング", exercise_time="30分以上",
                     exercise_frequency="１週間に２回以上",
                     exercise_intensity="少し汗をかく程度", daily_activity="1日5000歩以上",
                     other1="飲酒の制限、肥満度の改善", other2="毎日の歩数の測定"),
            Template(main_disease="脂質異常症", sheet_name="3_LDL70以下", goal1="LDLコレステロール＜100/TG＜150/HDL>40",
                     goal2="毎日の有酸素運動と食習慣の改善",
                     diet="脂肪の多い食品や甘い物を控える/食物繊維の摂取量を増やす/ゆっくり食べる/間食を減らす",
                     exercise_prescription="ウォーキング", exercise_time="30分以上",
                     exercise_frequency="１週間に２回以上",
                     exercise_intensity="少し汗をかく程度", daily_activity="1日5000歩以上",
                     other1="飲酒の制限、肥満度の改善", other2="毎日の歩数の測定"),
            Template(main_disease="糖尿病", sheet_name="1_HbA1c７％", goal1="HbA1ｃ７％/体重を当初の－３Kgとする",
                     goal2="1日5000歩以上の歩行/間食の制限/糖質の制限",
                     diet="食事量を適正にする/食物繊維の摂取量を増やす/ゆっくり食べる/間食を減らす",
                     exercise_prescription="ウォーキング", exercise_time="30分以上",
                     exercise_frequency="1週間に5回以上",
                     exercise_intensity="少し汗をかく程度", daily_activity="1日5000歩以上",
                     other1="睡眠の確保１日７時間", other2="毎日の歩数の測定"),
            Template(main_disease="糖尿病", sheet_name="2_HbA1c６％", goal1="HbA1ｃを正常化/HbA1ｃ6％",
                     goal2="１日５０００歩以上の歩行/間食の制限/糖質の制限",
                     diet="食事量を適正にする/食物繊維の摂取量を増やす/ゆっくり食べる/間食を減らす",
                     exercise_prescription="ウォーキング", exercise_time="30分以上",
                     exercise_frequency="１週間に５回以上",
                     exercise_intensity="少し汗をかく程度", daily_activity="1日5000歩以上",
                     other1="睡眠の確保１日７時間", other2="毎日の歩数の測定"),
            Template(main_disease="糖尿病", sheet_name="3_HbA1c８％", goal1="HbA1ｃを低血糖に注意して下げる",
                     goal2="ストレッチを中心とした運動/間食の制限/糖質の制限",
                     diet="食事量を適正にする/食物繊維の摂取量を増やす/ゆっくり食べる/間食を減らす",
                     exercise_prescription="ストレッチ運動", exercise_time="10分以上",
                     exercise_frequency="１週間に２回以上",
                     exercise_intensity="息切れしない程度", daily_activity="ストレッチ運動を主に行う",
                     other1="睡眠の確保１日７時間", other2="家庭での血圧の測定"),
        ]
        session.add_all(templates)
        session.commit()

    session.close()

    def on_startup(e):
        error_message_start, df_patients_data = load_patient_data()
        if error_message_start:
            snack_bar = ft.SnackBar(
                content=ft.Text(error_message_start),
                duration=1000
            )
            snack_bar.open = True
            page.overlay.append(snack_bar)
            page.update()

    def open_settings_dialog(e):
        def close_dialog(e):
            dialog.open = False
            page.update()

        def csv_export(e):
            export_to_csv(e)
            close_dialog(e)

        def csv_import(e):
            file_picker.pick_files(allow_multiple=False)
            close_dialog(e)

        content = ft.Container(
            content=ft.Column([
                ft.Text(f"LDTPapp\nバージョン: {VERSION}\n最終更新日: {LAST_UPDATED}"),
                ft.ElevatedButton("CSV出力", on_click=csv_export),
                ft.ElevatedButton("CSV取込", on_click=lambda _: file_picker.pick_files(allow_multiple=False)),
            ]),
            height=page.window.height * 0.3,
        )

        dialog = ft.AlertDialog(
            title=ft.Text("設定"),
            content=content,
            actions=[
                ft.TextButton("閉じる", on_click=close_dialog)
            ]
        )

        page.overlay.append(dialog)
        dialog.open = True
        page.update()

    # 設定ボタンを作成
    settings_button = ft.ElevatedButton("設定", on_click=open_settings_dialog)

    def on_file_selected(e: ft.FilePickerResultEvent):
        if e.files:
            file_path = e.files[0].path
            import_csv(file_path)

    file_picker = ft.FilePicker(on_result=on_file_selected)
    page.overlay.append(file_picker)

    def import_csv(file_path):
        file_name = os.path.basename(file_path)
        if not re.match(r'^patient_info_.*\.csv$', file_name):
            error_snack_bar = ft.SnackBar(
                content=ft.Text("インポートエラー:このファイルはインポートできません"),
                duration=2000
            )
            error_snack_bar.open = True
            page.overlay.append(error_snack_bar)
            page.update()
            return

        try:
            with open(file_path, 'r', encoding='shift_jis') as csvfile:
                csv_reader = csv.DictReader(csvfile)
                session = Session()
                for row in csv_reader:
                    patient_info = PatientInfo(
                        patient_id=int(row['patient_id']),
                        patient_name=row['patient_name'],
                        kana=row['kana'],
                        gender=row['gender'],
                        birthdate=datetime.strptime(row['birthdate'], '%Y-%m-%d').date(),
                        issue_date=datetime.strptime(row['issue_date'], '%Y-%m-%d').date(),
                        doctor_id=int(row['doctor_id']),
                        doctor_name=row['doctor_name'],
                        department=row['department'],
                        department_id=int(row['department_id']),
                        main_diagnosis=row['main_diagnosis'],
                        sheet_name=row['sheet_name'],
                        creation_count=int(row['creation_count']),
                        goal1=row['goal1'],
                        goal2=row['goal2'],
                        target_weight=float(row['target_weight']) if row['target_weight'] else None,
                        diet=row['diet'],
                        exercise_prescription=row['exercise_prescription'],
                        exercise_time=row['exercise_time'],
                        exercise_frequency=row['exercise_frequency'],
                        exercise_intensity=row['exercise_intensity'],
                        daily_activity=row['daily_activity'],
                        nonsmoker=row['nonsmoker'] == 'True',
                        smoking_cessation=row['smoking_cessation'] == 'True',
                        other1=row['other1'],
                        other2=row['other2']
                    )
                    session.add(patient_info)
                session.commit()
                session.close()

            snack_bar = ft.SnackBar(
                content=ft.Text("CSVファイルからデータがインポートされました"),
                duration=1000
            )
            snack_bar.open = True
            page.overlay.append(snack_bar)
            update_history(int(patient_id.value))
            page.update()

        except Exception as e:
            error_snack_bar = ft.SnackBar(
                content=ft.Text(f"インポート中にエラーが発生しました: {str(e)}"),
                duration=3000
            )
            error_snack_bar.open = True
            page.overlay.append(error_snack_bar)
            page.update()

    def export_to_csv(e):
        try:
            # CSVファイル名を現在の日時で生成
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            csv_filename = f"patient_info_export_{timestamp}.csv"
            csv_path = os.path.join(export_folder, csv_filename)

            # エクスポートフォルダが存在しない場合は作成
            os.makedirs(export_folder, exist_ok=True)

            # セッションを開始
            session = Session()

            # PatientInfoテーブルからすべてのデータを取得
            patient_data = session.query(PatientInfo).all()

            # CSVファイルを書き込みモードで開く
            with open(csv_path, 'w', newline='', encoding='shift_jis', errors='ignore') as csvfile:
                writer = csv.writer(csvfile)

                # ヘッダー行を書き込む
                writer.writerow([column.name for column in PatientInfo.__table__.columns])

                # データ行を書き込む
                for patient in patient_data:
                    writer.writerow([getattr(patient, column.name) for column in PatientInfo.__table__.columns])

            session.close()

            snack_bar = ft.SnackBar(
                content=ft.Text(f"データがCSVファイル '{csv_filename}' にエクスポートされました"),
                duration=1000
            )
            snack_bar.open = True
            page.overlay.append(snack_bar)
            page.update()

            os.startfile(export_folder)

        except Exception as e:
            # エラーメッセージを表示
            error_snack_bar = ft.SnackBar(
                content=ft.Text(f"エクスポート中にエラーが発生しました: {str(e)}"),
                duration=1000
            )
            error_snack_bar.open = True
            page.overlay.append(error_snack_bar)
            page.update()

    def on_issue_date_change(e):
        if issue_date_picker.value:
            issue_date_value.value = issue_date_picker.value.strftime("%Y/%m/%d")
            page.update()

    def on_main_diagnosis_change(e):
        selected_main_disease = main_diagnosis.value
        apply_template()

        with Session() as session:
            if selected_main_disease:
                main_disease = session.query(MainDisease).filter_by(
                    name=selected_main_disease).first()
                sheet_name_options = load_sheet_names(main_disease.id) if main_disease else []
            else:
                sheet_name_options = load_sheet_names(None)

        sheet_name_dropdown.options = sheet_name_options
        sheet_name_dropdown.value = ""
        page.update()

    def on_sheet_name_change(e):
        apply_template()
        page.update()

    def load_patient_info(patient_id_arg):
        patient_info = df_patients[df_patients.iloc[:, 2] == patient_id_arg]
        if not patient_info.empty:
            patient_info = patient_info.iloc[0]
            patient_id.value = str(patient_id_arg)
            issue_date_value.value = datetime.now().date().strftime("%Y/%m/%d")
            name_value.value = patient_info.iloc[3]
            kana_value.value = patient_info.iloc[4]
            gender_value.value = "男性" if patient_info.iloc[5] == 1 else "女性"
            birthdate = patient_info.iloc[6]
            birthdate_value.value = format_date(birthdate)
            doctor_id_value.value = str(patient_info.iloc[9])
            doctor_name_value.value = patient_info.iloc[10]
            department_value.value = patient_info.iloc[14]
            department_id_value.value = str(patient_info.iloc[13])
        else:
            # patient_infoが空の場合は空文字列を設定
            issue_date_value.value = ""
            name_value.value = ""
            kana_value.value = ""
            gender_value.value = ""
            birthdate_value.value = ""
            doctor_id_value.value = ""
            doctor_name_value.value = ""
            department_value.value = ""
        page.update()

    def create_treatment_plan_object(p_id, doctor_id, doctor_name, department, department_id, patients_df):
        patient_info_csv = patients_df.loc[patients_df.iloc[:, 2] == p_id]
        if patient_info_csv.empty:
            raise ValueError(f"患者ID {p_id} が見つかりません。")
        patient_info = patient_info_csv.iloc[0]
        return PatientInfo(
            patient_id=p_id,
            patient_name=patient_info.iloc[3],
            kana=patient_info.iloc[4],
            gender="男性" if patient_info.iloc[5] == 1 else "女性",
            birthdate=patient_info.iloc[6],
            issue_date=datetime.strptime(issue_date_value.value, "%Y/%m/%d").date(),
            doctor_id=doctor_id,
            doctor_name=doctor_name,
            department=department,
            department_id=department_id,
            main_diagnosis=main_diagnosis.value,
            sheet_name=sheet_name_dropdown.value,
            creation_count=1,
            target_weight=float(target_weight.value) if target_weight.value else None,
            goal1=goal1.value,
            goal2=goal2.value,
            diet=diet.value,
            exercise_prescription=exercise_prescription.value,
            exercise_time=exercise_time.value,
            exercise_frequency=exercise_frequency.value,
            exercise_intensity=exercise_intensity.value,
            daily_activity=daily_activity.value,
            nonsmoker=nonsmoker.value,
            smoking_cessation=smoking_cessation.value,
            other1=other1.value,
            other2=other2.value
        )

    def create_treatment_plan(p_id, doctor_id, doctor_name, department, department_id, patients_df):
        session = Session()
        try:
            treatment_plan = create_treatment_plan_object(p_id, doctor_id, doctor_name, department, department_id,
                                                          patients_df)
            session.add(treatment_plan)
            session.commit()
            TreatmentPlanGenerator.generate_plan(treatment_plan, "LDTPform")
            open_route(None)
        finally:
            session.close()

    def save_treatment_plan(p_id, doctor_id, doctor_name, department, department_id, patients_df):
        session = Session()
        try:
            treatment_plan = create_treatment_plan_object(p_id, doctor_id, doctor_name, department, department_id,
                                                          patients_df)
            session.add(treatment_plan)
            session.commit()
            open_route(None)
        finally:
            session.close()

    def show_error_message(message):
        snack_bar = ft.SnackBar(content=ft.Text(message), duration=1000)
        snack_bar.open = True
        page.overlay.append(snack_bar)
        page.update()

    def check_required_fields():
        if not main_diagnosis.value:
            show_error_message("主病名を選択してください")
            return False
        if not sheet_name_dropdown.value:
            show_error_message("シート名を選択してください")
            return False
        return True

    def create_new_plan(e):
        if not check_required_fields():
            return
        p_id = patient_id.value
        doctor_id = doctor_id_value.value
        doctor_name = doctor_name_value.value
        department_id = department_id_value.value
        department = department_value.value
        create_treatment_plan(int(p_id), int(doctor_id), doctor_name, department, int(department_id), df_patients)

    def save_new_plan(e):
        if not check_required_fields():
            return
        p_id = patient_id.value
        doctor_id = doctor_id_value.value
        doctor_name = doctor_name_value.value
        department_id = department_id_value.value
        department = department_value.value
        save_treatment_plan(int(p_id), int(doctor_id), doctor_name, department, int(department_id), df_patients)

    def print_plan(e):
        global selected_row
        session = Session()
        if selected_row is not None:
            patient_info = session.query(PatientInfo).filter(PatientInfo.id == selected_row['id']).first()
            if patient_info:
                patient_info.main_diagnosis = main_diagnosis.value
                patient_info.sheet_name = sheet_name_dropdown.value
                patient_info.creation_count = creation_count.value
                patient_info.target_weight = target_weight.value if target_weight.value else None
                patient_info.goal1 = goal1.value
                patient_info.goal2 = goal2.value
                patient_info.diet = diet.value
                patient_info.exercise_prescription = exercise_prescription.value
                patient_info.exercise_time = exercise_time.value
                patient_info.exercise_frequency = exercise_frequency.value
                patient_info.exercise_intensity = exercise_intensity.value
                patient_info.daily_activity = daily_activity.value
                patient_info.nonsmoker = nonsmoker.value
                patient_info.smoking_cessation = smoking_cessation.value
                patient_info.other1 = other1.value
                patient_info.other2 = other2.value
                session.commit()

                # 更新後のデータを使用して印刷
                TreatmentPlanGenerator.generate_plan(patient_info, "LDTPform")
        session.close()

    def on_patient_id_change(e):
        p_id = patient_id.value.strip()
        if patient_id:
            load_patient_info(int(p_id))
        update_history(p_id)

    def save_data(e):
        global selected_row
        session = Session()

        if selected_row is not None and 'id' in selected_row:
            patient_info = session.query(PatientInfo).filter(PatientInfo.id == selected_row['id']).first()
            if patient_info:
                if not check_required_fields():
                    return
                patient_info.patient_id = int(patient_id.value)
                patient_info.patient_name = name_value.value
                patient_info.kana = kana_value.value
                patient_info.gender = gender_value.value
                patient_info.birthdate = datetime.strptime(birthdate_value.value, "%Y/%m/%d").date()
                patient_info.issue_date = datetime.strptime(issue_date_value.value, "%Y/%m/%d").date()
                patient_info.doctor_id = int(doctor_id_value.value)
                patient_info.doctor_name = doctor_name_value.value
                patient_info.department = department_value.value
                patient_info.main_diagnosis = main_diagnosis.value
                patient_info.sheet_name = sheet_name_dropdown.value
                patient_info.creation_count = creation_count.value
                patient_info.target_weight = target_weight.value if target_weight.value else None
                patient_info.goal1 = goal1.value
                patient_info.goal2 = goal2.value
                patient_info.diet = diet.value
                patient_info.exercise_prescription = exercise_prescription.value
                patient_info.exercise_time = exercise_time.value
                patient_info.exercise_frequency = exercise_frequency.value
                patient_info.exercise_intensity = exercise_intensity.value
                patient_info.daily_activity = daily_activity.value
                patient_info.nonsmoker = nonsmoker.value
                patient_info.smoking_cessation = smoking_cessation.value
                patient_info.other1 = other1.value
                patient_info.other2 = other2.value
                session.commit()

                snack_bar = ft.SnackBar(ft.Text("データが保存されました"), duration=1000)
                snack_bar.open = True
                page.overlay.append(snack_bar)

            session.add(patient_info)
            session.commit()

        session.close()
        page.update()

    def copy_data(e):
        session = Session()
        patient_info = session.query(PatientInfo). \
            filter(PatientInfo.patient_id == patient_id.value). \
            order_by(PatientInfo.id.desc()).first()

        if patient_info:
            # pat.csvから最新の情報を取得
            error_message, df_patients = load_patient_data()
            if error_message:
                session.close()
                return

            patient_csv_info = df_patients[df_patients.iloc[:, 2] == int(patient_id.value)]
            if patient_csv_info.empty:
                session.close()
                return

            patient_csv_info = patient_csv_info.iloc[0]

            patient_info_copy = PatientInfo(
                patient_id=patient_info.patient_id,
                patient_name=patient_info.patient_name,
                kana=patient_info.kana,
                gender=patient_info.gender,
                birthdate=patient_info.birthdate,
                issue_date=datetime.now().date(),
                doctor_id=int(patient_csv_info.iloc[9]),
                doctor_name=patient_csv_info.iloc[10],
                department_id=int(patient_csv_info.iloc[13]),
                department=patient_csv_info.iloc[14],
                main_diagnosis=patient_info.main_diagnosis,
                sheet_name=patient_info.sheet_name,
                creation_count=patient_info.creation_count + 1,
                target_weight=patient_info.target_weight,
                goal1=patient_info.goal1,
                goal2=patient_info.goal2,
                diet=patient_info.diet,
                exercise_prescription=patient_info.exercise_prescription,
                exercise_time=patient_info.exercise_time,
                exercise_frequency=patient_info.exercise_frequency,
                exercise_intensity=patient_info.exercise_intensity,
                daily_activity=patient_info.daily_activity,
                nonsmoker=patient_info.nonsmoker,
                smoking_cessation=patient_info.smoking_cessation,
                other1=patient_info.other1,
                other2=patient_info.other2
            )
            session.add(patient_info_copy)
            session.commit()

            snack_bar = ft.SnackBar(
                ft.Text("前回データをコピーしました"),
                duration=1000,
            )
            snack_bar.open = True
            page.overlay.append(snack_bar)

        session.close()
        update_history(int(patient_id.value))
        page.update()

    def delete_data(e):
        session = Session()
        patient_info = session.query(PatientInfo).order_by(PatientInfo.id.desc()).first()
        if patient_info:
            session.delete(patient_info)
            session.commit()
            snack_bar = ft.SnackBar(
                ft.Text("データが削除されました"),
                duration=1000,
            )
            snack_bar.open = True
            page.overlay.append(snack_bar)

        session.close()
        open_route(e)

    def filter_data(e):
        update_history(patient_id.value)

    def update_history(filter_patient_id=None):
        data = fetch_data(filter_patient_id)
        history.rows = create_data_rows(data)
        page.update()

    def on_row_selected(e):
        global selected_row
        if e.data == "true":
            row_index = history.rows.index(e.control)
            selected_row = history.rows[row_index].data
            session = Session()
            patient_info = session.query(PatientInfo).filter(PatientInfo.id == selected_row['id']).first()
            if patient_info:
                patient_id.value = patient_info.patient_id

                # 発行日の更新
                issue_date_value.value = patient_info.issue_date.strftime("%Y/%m/%d") if patient_info.issue_date else ""

                # 主病名の更新
                main_diagnosis.options = load_main_diseases()
                main_diagnosis.value = patient_info.main_diagnosis

                # シート名の更新
                main_disease = session.query(MainDisease).filter_by(name=patient_info.main_diagnosis).first()
                if main_disease:
                    sheet_name_dropdown.options = load_sheet_names(main_disease.id)
                else:
                    sheet_name_dropdown.options = load_sheet_names(None)
                sheet_name_dropdown.value = patient_info.sheet_name

                creation_count.value = patient_info.creation_count
                target_weight.value = patient_info.target_weight
                goal1.value = patient_info.goal1
                goal2.value = patient_info.goal2
                diet.value = patient_info.diet
                exercise_prescription.value = patient_info.exercise_prescription
                exercise_time.value = patient_info.exercise_time
                exercise_frequency.value = patient_info.exercise_frequency
                exercise_intensity.value = patient_info.exercise_intensity
                daily_activity.value = patient_info.daily_activity
                nonsmoker.value = patient_info.nonsmoker
                smoking_cessation.value = patient_info.smoking_cessation
                other1.value = patient_info.other1
                other2.value = patient_info.other2
            session.close()
            page.update()

        if e.data == "true":
            row_index = history.rows.index(e.control)
            selected_row = history.rows[row_index].data
            open_edit(e)

        if e.data == "true":
            row_index = history.rows.index(e.control)
            selected_row = history.rows[row_index].data
            open_edit(e)

    def fetch_data(filter_patient_id=None):
        if not filter_patient_id:
            return []

        session_fetch_data = Session()
        query = session_fetch_data.query(PatientInfo.id, PatientInfo.issue_date, PatientInfo.department,
                                         PatientInfo.doctor_name, PatientInfo.main_diagnosis,
                                         PatientInfo.sheet_name, PatientInfo.creation_count). \
            order_by(PatientInfo.patient_id.asc(), PatientInfo.id.desc())

        query = query.filter(PatientInfo.patient_id == filter_patient_id)

        return ({
            "id": str(info.id),
            "issue_date": info.issue_date.strftime("%Y/%m/%d") if info.issue_date else "",
            "department": info.department,
            "doctor_name": info.doctor_name,
            "main_diagnosis": info.main_diagnosis,
            "sheet_name": info.sheet_name,
            "count": info.creation_count
        } for info in query)

    def create_data_rows(data):
        rows = []
        for item in data:
            row = ft.DataRow(
                cells=[
                    ft.DataCell(ft.Text(item["issue_date"])),
                    ft.DataCell(ft.Text(item["department"])),
                    ft.DataCell(ft.Text(item["doctor_name"])),
                    ft.DataCell(ft.Text(item["main_diagnosis"])),
                    ft.DataCell(ft.Text(item["sheet_name"])),
                    ft.DataCell(ft.Text(item["count"])),
                ],
                on_select_changed=on_row_selected,
                data=item
            )
            rows.append(row)
        return rows

    def apply_template(e=None):
        session_apply_template = Session()
        try:
            template = session_apply_template.query(Template).filter(
                Template.main_disease == main_diagnosis.value,
                Template.sheet_name == sheet_name_dropdown.value
            ).first()
            if template:
                goal1.value = template.goal1
                goal2.value = template.goal2
                diet.value = template.diet
                exercise_prescription.value = template.exercise_prescription
                exercise_time.value = template.exercise_time
                exercise_frequency.value = template.exercise_frequency
                exercise_intensity.value = template.exercise_intensity
                daily_activity.value = template.daily_activity
                other1.value = template.other1
                other2.value = template.other2
        finally:
            session_apply_template.close()
        page.update()

    def save_template(e):
        if not main_diagnosis.value:
            snack_bar = ft.SnackBar(content=ft.Text("主病名を選択してください"))
            snack_bar.open = True
            page.overlay.append(snack_bar)
            page.update()
            return
        if not sheet_name_dropdown.value:
            snack_bar = ft.SnackBar(content=ft.Text("シート名を選択してください"))
            snack_bar.open = True
            page.overlay.append(snack_bar)
            page.update()
            return

        session = Session()
        template = session.query(Template).filter(Template.main_disease == main_diagnosis.value,
                                                  Template.sheet_name == sheet_name_dropdown.value).first()
        if template:
            template.goal1 = goal1.value
            template.goal2 = goal2.value
            template.diet = diet.value
            template.exercise_prescription = exercise_prescription.value
            template.exercise_time = exercise_time.value
            template.exercise_frequency = exercise_frequency.value
            template.exercise_intensity = exercise_intensity.value
            template.daily_activity = daily_activity.value
            template.other1 = other1.value
            template.other2 = other2.value
        else:
            template = Template(
                main_disease=main_diagnosis.value,
                sheet_name=sheet_name_dropdown.value,
                goal1=goal1.value,
                goal2=goal2.value,
                diet=diet.value,
                exercise_prescription=exercise_prescription.value,
                exercise_time=exercise_time.value,
                exercise_frequency=exercise_frequency.value,
                exercise_intensity=exercise_intensity.value,
                daily_activity=daily_activity.value,
                other1=other1.value,
                other2=other2.value
            )
            session.add(template)
        session.commit()
        session.close()

        snack_bar = ft.SnackBar(
            content=ft.Text("テンプレートが保存されました"),
            duration=1000)
        snack_bar.open = True
        page.overlay.append(snack_bar)
        page.update()
        open_route(None)

    def route_change(e):
        page.views.clear()
        page.views.append(
            View(
                "/",
                [
                    ft.Row(
                        controls=[
                            patient_id,
                            name_value,
                            kana_value,
                            gender_value,
                            birthdate_value,
                        ]
                    ),
                    ft.Row(
                        controls=[
                            doctor_id_value,
                            doctor_name_value,
                            department_id_value,
                            department_value,
                            settings_button,
        ]
                    ),
                    ft.Row(
                        controls=[
                            buttons,
                            ft.Text("(SOAP画面を閉じるとアプリが終了します)", size=12)
                        ]
                    ),
                    ft.Row(
                        controls=[
                            ft.Text("計画書一覧", size=16),
                            ft.Text("計画書をクリックすると編集画面が表示されます", size=14),
                        ]
                    ),
                    ft.Divider(),
                    history_scrollable,
                ],
            )
        )

        if page.route == "/create":
            page.views.append(
                View(
                    "/create",
                    [
                        ft.Row(
                            controls=[
                                ft.Container(
                                    content=ft.Text("新規作成", size=16, weight=ft.FontWeight.BOLD),
                                    border=ft.border.all(3, ft.colors.BLUE),
                                    padding=5,
                                    border_radius=5,
                                ),
                                main_diagnosis,
                                sheet_name_dropdown,
                                creation_count,
                                ft.Text("回目", size=14),
                                issue_date_row,
                            ]
                        ),
                        ft.Row(
                            controls=[
                                goal1,
                                target_weight,
                                ft.Text("kg", size=14),
                            ]
                        ),
                        goal2,
                        guidance_items,
                        create_buttons,
                    ],
                )
            )
        page.update()

        if page.route == "/edit":
            page.views.append(
                View(
                    "/edit",
                    [
                        ft.Row(
                            controls=[
                                ft.Container(
                                    content=ft.Text("編集", size=16, weight=ft.FontWeight.BOLD),
                                    border=ft.border.all(3, ft.colors.BLUE),
                                    padding=5,
                                    border_radius=5,
                                ),
                                main_diagnosis,
                                sheet_name_dropdown,
                                creation_count,
                                ft.Text("回目", size=14),
                                issue_date_row,
                            ]
                        ),
                        ft.Row(
                            controls=[
                                goal1,
                                target_weight,
                                ft.Text("kg", size=14),
                            ]
                        ),
                        goal2,
                        guidance_items,
                        edit_buttons,
                    ],
                )
            )
        page.update()

        if page.route == "/template":
            page.views.append(
                View(
                    "/template",
                    [
                        ft.Row(
                            controls=[
                                ft.Container(
                                    content=ft.Text("テンプレート", size=16, weight=ft.FontWeight.BOLD),
                                    border=ft.border.all(3, ft.colors.BLUE),
                                    padding=5,
                                    border_radius=5,
                                ),
                                main_diagnosis,
                                sheet_name_dropdown,
                            ]
                        ),
                        ft.Row(
                            controls=[
                                goal1,
                            ]
                        ),
                        goal2,
                        guidance_items_template,
                        template_buttons,
                    ],
                )
            )
        page.update()

    def view_pop(e):
        page.views.pop()
        top_view = page.views[-1]
        page.go(top_view.route)

    def open_create(e):
        page.go("/create")

    def open_edit(e):
        page.go("/edit")

    def open_template(e):
        page.go("/template")

    def open_route(e):
        for field in [target_weight, goal1, goal2, diet,
                      exercise_prescription, exercise_time, exercise_frequency, exercise_intensity,
                      daily_activity, other1, other2]:
            field.value = ""

        main_diagnosis.value = ""
        sheet_name_dropdown.value = ""
        creation_count.value = 1  # 作成回数の初期値を再設定
        nonsmoker.value = False
        smoking_cessation.value = False

        page.go("/")
        update_history(int(patient_id.value))
        page.update()

    def on_close(e):
        page.window.close()

    # Patient Information
    patient_id = ft.TextField(label="患者ID", on_change=on_patient_id_change, value=initial_patient_id, width=150)
    issue_date_value = ft.TextField(label="発行日", width=150, read_only=True)

    issue_date_picker = ft.DatePicker(
        on_change=on_issue_date_change,
    )
    page.overlay.append(issue_date_picker)

    issue_date_button = ft.ElevatedButton(
        "日付選択",
        icon=ft.icons.CALENDAR_TODAY,
        on_click=lambda _: issue_date_picker.pick_date()
    )

    # 発行日の行を作成
    issue_date_row = ft.Row([issue_date_value, issue_date_button])

    name_value = ft.TextField(label="氏名", read_only=True, width=150)
    kana_value = ft.TextField(label="カナ", read_only=True, width=150)
    gender_value = ft.TextField(label="性別", read_only=True, width=150)
    birthdate_value = ft.TextField(label="生年月日", read_only=True, width=150)
    doctor_id_value = ft.TextField(label="医師ID", read_only=True, width=150)
    doctor_name_value = ft.TextField(label="医師名", read_only=True, width=150)
    department_id_value = ft.TextField(label="診療科ID", read_only=True, width=150)
    department_value = ft.TextField(label="診療科", read_only=True, width=150)
    main_disease_options = load_main_diseases()
    main_diagnosis = ft.Dropdown(label="主病名", options=main_disease_options, width=200, text_size=14, value="",
                                 on_change=on_main_diagnosis_change, autofocus=True)
    sheet_name_options = load_sheet_names(main_diagnosis.value)
    sheet_name_dropdown = ft.Dropdown(label="シート名", options=sheet_name_options, width=300, text_size=14, value="",
                                      on_change=on_sheet_name_change)
    creation_count = ft.TextField(
        label="作成回数",
        width=100,
        value="1",
        on_submit=lambda _: goal1.focus(),
    )
    target_weight = ft.TextField(label="目標体重", width=100, value="", on_submit=lambda _: goal2.focus())
    goal1 = ft.TextField(label="①達成目標：患者と相談した目標", width=700, value="達成目標を入力してください",
                         on_submit=lambda _: target_weight.focus())
    goal2 = ft.TextField(label="②行動目標：患者と相談した目標", width=700, value="行動目標を入力してください",
                         on_submit=lambda _: diet.focus())

    diet = ft.TextField(
        label="食事",
        multiline=False,
        disabled=False,
        value="",
        width=1000,
        on_submit=lambda _: exercise_prescription.focus()
    )

    exercise_prescription, exercise_time, exercise_frequency, exercise_intensity, daily_activity = create_form_fields(dropdown_items)

    def on_tobacco_checkbox_change(e):
        if e.control == nonsmoker and nonsmoker.value:
            smoking_cessation.value = False
            smoking_cessation.update()
        elif e.control == smoking_cessation and smoking_cessation.value:
            nonsmoker.value = False
            nonsmoker.update()

    nonsmoker = ft.Checkbox(label="非喫煙者である", on_change=on_tobacco_checkbox_change)
    smoking_cessation = ft.Checkbox(label="禁煙の実施方法等を指示", on_change=on_tobacco_checkbox_change)

    other1 = ft.TextField(label="その他1", value="", width=400, on_submit=lambda _: other2.focus())
    other2 = ft.TextField(label="その他2", value="", width=400)

    guidance_items = ft.Column([
        diet,
        ft.Row([exercise_prescription, exercise_time, exercise_frequency, exercise_intensity]),
        daily_activity,
        ft.Row([ft.Text("たばこ", size=14), nonsmoker, smoking_cessation,
                ft.Text("    (チェックボックスを2回選ぶと解除できます)", size=12)]),
        ft.Row([other1, other2]),
    ])

    guidance_items_template = ft.Column([
        diet,
        ft.Row([exercise_prescription, exercise_time, exercise_frequency, exercise_intensity]),
        daily_activity,
        ft.Row([other1, other2]),
    ])

    selected_row = None
    data = fetch_data()
    rows = create_data_rows(data)

    history = ft.DataTable(
        columns=[
            ft.DataColumn(ft.Text("発行日")),
            ft.DataColumn(ft.Text("診療科")),
            ft.DataColumn(ft.Text("医師名")),
            ft.DataColumn(ft.Text("主病名")),
            ft.DataColumn(ft.Text("シート名")),
            ft.DataColumn(ft.Text("作成回数")),
        ],
        rows=rows,
        width=1200,
    )

    history_column = ft.Column([history], scroll=ft.ScrollMode.AUTO, width=1200, height=400)
    history_scrollable = ft.Container(
        content=history_column,
        width=1200,
        height=400,
        border=ft.border.all(1, ft.colors.BLACK),
        border_radius=5,
        padding=10,
    )

    buttons = ft.Row([
        ft.ElevatedButton("新規作成", on_click=open_create),
        ft.ElevatedButton("前回コピー", on_click=copy_data),
        ft.ElevatedButton("テンプレート編集", on_click=open_template),
        ft.ElevatedButton("閉じる", on_click=on_close),
    ])

    create_buttons = ft.Row([
        ft.ElevatedButton("新規登録して印刷", on_click=create_new_plan),
        ft.ElevatedButton("新規登録", on_click=save_new_plan),
        ft.ElevatedButton("戻る", on_click=open_route),
    ])

    edit_buttons = ft.Row([
        ft.ElevatedButton("保存", on_click=save_data),
        ft.ElevatedButton("印刷", on_click=print_plan),
        ft.ElevatedButton("削除", on_click=delete_data),
        ft.ElevatedButton("戻る", on_click=open_route),
    ])

    template_buttons = ft.Row([
        ft.ElevatedButton("保存", on_click=save_template),
        ft.ElevatedButton("戻る", on_click=open_route),
    ])

    # page画面の設定
    layout = ft.Column([
        ft.Row(
            controls=[]
        ),
    ])

    page.add(layout)
    update_history()

    if initial_patient_id:
        load_patient_info(int(initial_patient_id))
        patient_id.value = initial_patient_id
        filter_data(patient_id.value)
        update_history(patient_id.value)

    page.window.on_resized = on_startup
    page.on_route_change = route_change
    page.on_view_pop = view_pop
    page.go(page.route)


def main(page: ft.Page):
    start_file_monitoring(page)
    check_file_exists(page)
    create_ui(page)


if __name__ == "__main__":
    ft.app(target=main)
