import os
from datetime import datetime

import flet as ft
from flet import AppBar, ElevatedButton, Page, Text, View
import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import create_engine, Column, Integer, String, Float, Date
from sqlalchemy.orm import declarative_base
from sqlalchemy.orm import sessionmaker


# SQLAlchemyの設定
db_url = "sqlite:///ldtp_app.db"
engine = create_engine(db_url)

# engine = create_engine("postgresql+psycopg2://postgres:postgres@localhost/fletapp")
# engine = create_engine("postgresql+psycopg2://postgres:postgres@192.168.3.5/fletapp")

Session = sessionmaker(bind=engine)

Base = declarative_base()


# モデルの定義
class TreatmentPlan(Base):
    __tablename__ = "treatment_plans"
    id = Column(Integer, primary_key=True)
    patient_id = Column(Integer)
    issue_date = Column(Date)
    doctor_id = Column(Integer)
    doctor_name = Column(String)
    department = Column(String)
    creation_count = Column(Integer)
    main_disease = Column(String)
    sheet_name = Column(String)
    weight = Column(Float)
    file_path = Column(String)


class MainDisease(Base):
    __tablename__ = "main_diseases"
    id = Column(Integer, primary_key=True)
    name = Column(String)


class SheetName(Base):
    __tablename__ = "sheet_names"
    id = Column(Integer, primary_key=True)
    name = Column(String)


# テーブルの作成
Base.metadata.create_all(engine)


def load_patient_data():
    date_columns = [0, 6]
    return pd.read_csv(r"C:\InnoKarte\pat.csv", encoding="shift_jis", header=None, parse_dates=date_columns)


def create_treatment_plan(patient_id, doctor_id, doctor_name, department, creation_count, main_disease, sheet_name, weight,
                          df_patients):
    session = Session()
    current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
    workbook = load_workbook(r"C:\Shinseikai\LDTPapp\生活習慣病療養計画書.xlsm", keep_vba=True)
    common_sheet = workbook["共通情報"]

    patient_info = df_patients.loc[df_patients.iloc[:, 2] == patient_id]

    if patient_info.empty:
        session.close()
        raise ValueError(f"患者ID {patient_id} が見つかりません。")

    patient_info = patient_info.iloc[0]

    # 共通情報シートに必要な情報を設定
    common_sheet["B2"] = patient_info.iloc[2]
    common_sheet["B3"] = patient_info.iloc[3]
    common_sheet["B4"] = patient_info.iloc[4]
    common_sheet["B5"] = "男性" if patient_info.iloc[5] == 1 else "女性"
    common_sheet["B6"] = patient_info.iloc[6]
    common_sheet["B8"] = doctor_name
    common_sheet["B1"] = datetime.now().strftime("%Y/%m/%d")
    common_sheet["B7"] = doctor_id
    common_sheet["B13"] = patient_info.iloc[10]
    common_sheet["B10"] = department
    common_sheet["B11"] = creation_count
    common_sheet["B12"] = main_disease
    common_sheet["B14"] = sheet_name
    common_sheet["b13"] = weight

    new_file_name = f"生活習慣病療養計画書_{current_datetime}.xlsm"
    file_path = r"C:\Shinseikai\LDTPapp" + "\\" + new_file_name
    workbook.save(file_path)
    wb = load_workbook(file_path, read_only=False, keep_vba=True)
    wb.active = wb[sheet_name]
    wb.save(file_path)
    os.startfile(file_path)

    treatment_plan = TreatmentPlan(
        patient_id=patient_id,
        issue_date=datetime.now().date(),
        doctor_id=doctor_id,
        department=department,
        creation_count=creation_count,
        main_disease=main_disease,
        sheet_name=sheet_name,
        weight=weight,
        file_path=file_path
    )
    session.add(treatment_plan)
    session.commit()
    session.close()


def get_issued_plans():
    session = Session()
    issued_plans = session.query(TreatmentPlan).all()
    session.close()
    return issued_plans


def load_main_diseases():
    session = Session()
    main_diseases = session.query(MainDisease).all()
    session.close()
    return [ft.dropdown.Option(disease.name) for disease in main_diseases]


def load_sheet_names():
    session = Session()
    sheet_names = session.query(SheetName).all()
    session.close()
    return [ft.dropdown.Option(sheet.name) for sheet in sheet_names]


def format_date(date_str):
    if pd.isna(date_str):
        return ""
    return pd.to_datetime(date_str).strftime("%Y/%m/%d")


def main(page: ft.Page):
    page.title = "生活習慣病療養計画書"
    page.window_width = 1000
    page.window_height = 800

    # 初期データの挿入
    session = Session()
    if session.query(MainDisease).count() == 0:
        main_diseases = [
            MainDisease(name="高血圧"),
            MainDisease(name="脂質異常症"),
            MainDisease(name="糖尿病")
        ]
        session.add_all(main_diseases)
        session.commit()

    if session.query(SheetName).count() == 0:
        sheet_names = [
            SheetName(name="DM01"),
            SheetName(name="DM02"),
            SheetName(name="DM03")
        ]
        session.add_all(sheet_names)
        session.commit()

    session.close()

    df_patients = load_patient_data()
    print(df_patients)
    # CSVファイルから1行目の患者IDを取得
    initial_patient_id = ""
    if not df_patients.empty:
        initial_patient_id = str(df_patients.iloc[0, 2])

    def load_patient_info(patient_id):
        patient_info = df_patients[df_patients.iloc[:, 2] == patient_id]
        if not patient_info.empty:
            patient_info = patient_info.iloc[0]
            issue_date_value.value = format_date(patient_info.iloc[0])
            name_value.value = patient_info.iloc[3]
            kana_value.value = patient_info.iloc[4]
            gender_value.value = "男性" if patient_info.iloc[5] == 1 else "女性"
            birthdate = patient_info.iloc[6]
            birthdate_value.value = format_date(birthdate)
            doctor_id_value.value = str(patient_info.iloc[9])
            doctor_name_value.value = patient_info.iloc[10]
            department_value.value = patient_info.iloc[14]
        else:
            # patient_infoが空の場合は空文字列を設定
            issue_date_value.value = ""
            name_value.value = ""
            kana_value.value = ""
            gender_value.value = ""
            birthdate_value.value = ""
            doctor_id_value.value = ""
            doctor_name_value.value = ""
            department_value.value = ""
        page.update()

    def create_new_plan(e):
        patient_id = patient_id_value.value.strip()
        doctor_id = doctor_id_value.value.strip()
        doctor_name = doctor_name_value.value
        if not patient_id or not doctor_id:
            page.snack_bar = ft.SnackBar(content=ft.Text("患者IDと医師IDは必須です"))
            page.snack_bar.open = True
            page.update()
            return
        department = department_value.value
        creation_count = int(creation_count_value.value)
        main_disease = main_disease_dropdown.value
        sheet_name = sheet_name_dropdown.value
        weight = float(weight_value.value)

        create_treatment_plan(int(patient_id), int(doctor_id), doctor_name, department, creation_count, main_disease, sheet_name,
                              weight, df_patients)
        page.snack_bar = ft.SnackBar(content=ft.Text("計画書ファイルが作成されました"))
        view_issued_plans(None)
        page.update()

    def view_issued_plans(e):
        issued_plans = get_issued_plans()
        plans_data = []
        for plan in issued_plans:
            plans_data.append([
                plan.id,
                str(plan.issue_date),
                str(plan.patient_id),
                plan.department,
                plan.main_disease,
                plan.sheet_name,
                str(plan.creation_count),
            ])
        issued_plans_table.rows = [ft.DataRow(cells=[ft.DataCell(ft.Text(cell)) for cell in row]) for row in plans_data]
        page.update()

    def print_plan(e):
        pass

    def delete_plan(e):
        if not selected_plan_id:
            page.snack_bar = ft.SnackBar(content=ft.Text("削除する計画書を選択してください"))
            page.snack_bar.open = True
            page.update()
            return
        session = Session()
        plan = session.query(TreatmentPlan).filter_by(id=selected_plan_id).first()
        if plan:
            session.delete(plan)
            session.commit()
            view_issued_plans(None)
            page.snack_bar = ft.SnackBar(content=ft.Text("選択した計画書を削除しました"))
            page.snack_bar.open = True
        else:
            page.snack_bar = ft.SnackBar(content=ft.Text("選択した計画書は見つかりませんでした"))
            page.snack_bar.open = True
        session.close()
        page.update()

    def on_patient_id_change(e):
        patient_id = patient_id_value.value.strip()
        if patient_id:
            load_patient_info(int(patient_id))

    def route_change(e):
        print("Route change:", e.route)
        page.views.clear()

        # トップページ（常にviewに追加する）
        page.views.append(
            View(
                "/",
                [
                    ft.Row(
                        controls=[
                            patient_id_value,
                            issue_date_value,
                            name_value,
                            kana_value,
                            gender_value,
                            birthdate_value
                        ]
                    ),
                    ft.Row(
                        controls=[
                            doctor_id_value,
                            doctor_name_value,
                            department_value,
                        ]
                    ),
                    ft.Row(
                        controls=[
                            main_disease_dropdown,
                            sheet_name_dropdown,
                            creation_count_value,
                            weight_value
                        ]
                    ),
                    ft.Row(
                        controls=[
                            create_button,
                            print_button,
                            delete_button,
                            close_button
                        ]
                    ),
                    ElevatedButton("テストページへ移動", on_click=open_test),
                    ft.Divider(),
                    issued_plans_table,
                ],
            )
        )
        # テストページ（テストページのときだけviewに追加する）
        if page.route == "/test":
            page.views.append(
                View(
                    "/test",
                    [
                        ft.Row(
                            controls=[
                                main_disease_dropdown,
                                sheet_name_dropdown,
                                creation_count_value,
                                weight_value
                            ]
                        ),
                        goal1,
                        goal2,
                        guidance_items,
                        buttons
                    ],
                )
            )

        # ページ更新
        page.update()

    # 現在のページを削除して、前のページに戻る
    def view_pop(e):
        print("View pop:", e.view)
        page.views.pop()
        top_view = page.views[-1]
        page.go(top_view.route)

    # テストページへ移動
    def open_test(e):
        page.go("/test")

    def route_test(e):
        page.go("/")

    patient_id_value = ft.TextField(label="患者ID", on_change=on_patient_id_change, value=initial_patient_id, width=150)
    issue_date_value = ft.TextField(label="発行日", read_only=True, width=150)
    name_value = ft.TextField(label="氏名", read_only=True, width=150)
    kana_value = ft.TextField(label="カナ", read_only=True, width=150)
    gender_value = ft.TextField(label="性別", read_only=True, width=150)
    birthdate_value = ft.TextField(label="生年月日", read_only=True, width=150)

    main_disease_options = load_main_diseases()
    sheet_name_options = load_sheet_names()

    doctor_id_value = ft.TextField(label="医師ID", read_only=True, width=150)
    doctor_name_value = ft.TextField(label="医師名", read_only=True, width=150)
    department_value = ft.TextField(label="診療科", read_only=True, width=150)
    creation_count_value = ft.TextField(label="作成回数", width=150)
    main_disease_dropdown = ft.Dropdown(label="主病名", options=main_disease_options, width=150)
    sheet_name_dropdown = ft.Dropdown(label="シート名", options=sheet_name_options, width=150)
    weight_value = ft.TextField(label="目標体重", width=150)

    create_button = ft.ElevatedButton("新規登録", on_click=create_new_plan)
    print_button = ft.ElevatedButton("印刷", on_click=print_plan)
    delete_button = ft.ElevatedButton("削除", on_click=delete_plan)
    close_button = ft.ElevatedButton("閉じる", on_click=lambda e: None )

    goal1 = ft.TextField(label="①達成目標：患者と相談した目標", width=600, value="")
    goal2 = ft.TextField(label="②行動目標：患者と相談した目標", width=600, value="")

    # Guidance Items
    guidance_items = ft.Column([
        ft.TextField(
            label="食事",
            multiline=True,
            disabled=False,
            value="食事量を適正にする\n食物繊維の摂取量を増やす\nゆっくり食べる\n間食を減らす",
            width=400,
        ),
        ft.TextField(label="運動処方", width=400, value="ウォーキング"),
        ft.Row([
            ft.TextField(label="時間", value="30分以上", width=200),
            ft.TextField(label="頻度", value="ほぼ毎日", width=200),
            ft.TextField(label="強度", value="少し汗をかく程度", width=200),
        ]),
        ft.TextField(label="日常生活の活動量増加", value="1日8000歩以上", width=400),
        ft.Text("たばこ", size=14),
        ft.Row([
            ft.Checkbox(label="非喫煙者である"),
            ft.Checkbox(label="禁煙の実施方法等を指示"),
        ]),
        ft.Row([
            ft.TextField(label="その他1", value="睡眠の確保１日７時間", width=300),
            ft.TextField(label="その他2", value="家庭での毎日の歩数の測定", width=300),
        ]),
    ])

    buttons = ft.Row([
        ft.ElevatedButton("戻る", on_click=route_test),
        ft.ElevatedButton("保存", on_click=lambda e: None),
    ])

    issued_plans_table = ft.DataTable(
        columns=[
            ft.DataColumn(ft.Text("ID")),
            ft.DataColumn(ft.Text("発行日")),
            ft.DataColumn(ft.Text("患者ID")),
            ft.DataColumn(ft.Text("診療科")),
            ft.DataColumn(ft.Text("主病名")),
            ft.DataColumn(ft.Text("シート名")),
            ft.DataColumn(ft.Text("作成回数")),
        ],
        rows=[],  # データは後で追加
    )

    # 初期患者情報を読み込む
    if initial_patient_id:
        load_patient_info(initial_patient_id)
        selected_plan_id = None
        view_issued_plans(None)

    # ページ遷移イベントが発生したら、ページを更新
    page.on_route_change = route_change
    # AppBarの戻るボタンクリック時、前のページへ戻る
    page.on_view_pop = view_pop

    page.go(page.route)


ft.app(target=main)